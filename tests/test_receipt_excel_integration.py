"""receipt_excel_generatorツールの統合テスト"""
import pytest
from pathlib import Path
from unittest.mock import MagicMock
from tools.excel_generator import receipt_excel_generator
from openpyxl import load_workbook


class TestReceiptExcelGeneratorIntegration:
    """receipt_excel_generatorツールの統合テスト"""
    
    def setup_method(self):
        """各テストの前に実行される初期化処理"""
        # モックのToolContextを作成
        self.mock_context = MagicMock()
        self.mock_context.invocation_state = {
            "applicant_name": "田中太郎",
            "application_date": "2025-01-15",
            "session_id": "test_session_123"
        }
    
    def test_valid_receipt_generation(self):
        """正常な経費精算申請書の生成テスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=1500.0,
            date="2025-01-15",
            items=["ボールペン", "ノート"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is True
        assert "file_path" in result
        assert result["file_path"] != ""
        assert "正常に作成しました" in result["message"]
        
        # ファイルが実際に作成されたか確認
        file_path = Path(result["file_path"])
        assert file_path.exists()
        
        # Excelファイルの内容を検証
        wb = load_workbook(file_path)
        ws = wb.active
        
        # タイトル確認
        assert ws["A1"].value == "経費精算申請書"
        
        # 申請者名確認
        assert ws["A3"].value == "申請者名"
        assert ws["B3"].value == "田中太郎"
        
        # 申請日確認
        assert ws["A4"].value == "申請日"
        assert ws["B4"].value == "2025-01-15"
        
        # 店舗名確認
        assert ws["A6"].value == "店舗名"
        assert ws["B6"].value == "コンビニA"
        
        # 金額確認
        assert ws["A7"].value == "金額"
        assert "1,500" in ws["B7"].value
        
        # 購入日確認
        assert ws["A8"].value == "購入日"
        assert ws["B8"].value == "2025-01-15"
        
        # 経費区分確認
        assert ws["A9"].value == "経費区分"
        assert ws["B9"].value == "事務用品費"
        
        # 品目確認
        assert ws["A10"].value == "品目"
        assert "ボールペン" in ws["B10"].value
        assert "ノート" in ws["B10"].value
        
        # 承認状況確認
        assert ws["A12"].value == "承認状況"
        assert ws["B12"].value == "承認待ち"
        
        wb.close()
        
        # テスト後のクリーンアップ
        if file_path.exists():
            file_path.unlink()
        
        print("✅ 正常な経費精算申請書が生成され、内容も正しい")
    
    def test_invalid_empty_store_name(self):
        """空の店舗名でのバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="",  # 空文字列
            amount=1500.0,
            date="2025-01-15",
            items=["ボールペン"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "store_name" in result["message"]
        print(f"✅ 空の店舗名を検出: {result['message']}")
    
    def test_invalid_zero_amount(self):
        """金額が0の場合のバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=0.0,  # 0は許可されない
            date="2025-01-15",
            items=["ボールペン"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "amount" in result["message"]
        print(f"✅ 金額0を検出: {result['message']}")
    
    def test_invalid_negative_amount(self):
        """負の金額の場合のバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=-1500.0,  # 負の金額は許可されない
            date="2025-01-15",
            items=["ボールペン"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "amount" in result["message"]
        print(f"✅ 負の金額を検出: {result['message']}")
    
    def test_invalid_date_format(self):
        """無効な日付形式のバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=1500.0,
            date="2025/01/15",  # スラッシュ区切りは無効
            items=["ボールペン"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "date" in result["message"]
        print(f"✅ 無効な日付形式を検出: {result['message']}")
    
    def test_invalid_empty_items_list(self):
        """空の品目リストのバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=1500.0,
            date="2025-01-15",
            items=[],  # 空リストは許可されない
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "items" in result["message"]
        print(f"✅ 空の品目リストを検出: {result['message']}")
    
    def test_invalid_empty_expense_category(self):
        """空の経費区分のバリデーションエラーテスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount=1500.0,
            date="2025-01-15",
            items=["ボールペン"],
            expense_category="",  # 空文字列は許可されない
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        assert "expense_category" in result["message"]
        print(f"✅ 空の経費区分を検出: {result['message']}")
    
    def test_multiple_validation_errors(self):
        """複数のバリデーションエラーを同時に検出するテスト"""
        result = receipt_excel_generator(
            store_name="",  # エラー1: 空の店舗名
            amount=-1500.0,  # エラー2: 負の金額
            date="2025/01/15",  # エラー3: 無効な日付形式
            items=[],  # エラー4: 空の品目リスト
            expense_category="",  # エラー5: 空の経費区分
            tool_context=self.mock_context
        )
        
        assert result["success"] is False
        assert result["file_path"] == ""
        assert "入力データが不正です" in result["message"]
        
        # 複数のエラーが含まれていることを確認
        message = result["message"]
        error_count = message.count(":")
        assert error_count >= 3  # 少なくとも3つのエラーが含まれている
        
        print(f"✅ 複数のバリデーションエラーを検出: {result['message']}")
    
    def test_multiple_items_generation(self):
        """複数の品目を含む申請書の生成テスト"""
        result = receipt_excel_generator(
            store_name="文房具店",
            amount=3500.0,
            date="2025-01-15",
            items=["ボールペン", "ノート", "クリアファイル", "付箋"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is True
        
        # ファイルの内容を検証
        file_path = Path(result["file_path"])
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 品目が全て含まれているか確認
        items_cell = ws["B10"].value
        assert "ボールペン" in items_cell
        assert "ノート" in items_cell
        assert "クリアファイル" in items_cell
        assert "付箋" in items_cell
        
        wb.close()
        
        # テスト後のクリーンアップ
        if file_path.exists():
            file_path.unlink()
        
        print("✅ 複数の品目を含む申請書が正常に生成された")
    
    def test_string_to_float_coercion(self):
        """文字列から数値への自動変換テスト"""
        result = receipt_excel_generator(
            store_name="コンビニA",
            amount="1500",  # 文字列で渡す
            date="2025-01-15",
            items=["ボールペン"],
            expense_category="事務用品費",
            tool_context=self.mock_context
        )
        
        assert result["success"] is True
        
        # ファイルの内容を検証
        file_path = Path(result["file_path"])
        wb = load_workbook(file_path)
        ws = wb.active
        
        # 金額が正しく表示されているか確認
        assert "1,500" in ws["B7"].value
        
        wb.close()
        
        # テスト後のクリーンアップ
        if file_path.exists():
            file_path.unlink()
        
        print("✅ 文字列から数値への自動変換が正常に機能")


if __name__ == "__main__":
    pytest.main([__file__, "-v", "-s"])
