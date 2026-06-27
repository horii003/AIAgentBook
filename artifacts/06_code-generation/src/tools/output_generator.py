"""申請書生成ツール

申請書テンプレート（Excel）に申請情報を埋め込み、申請書ファイルを生成する。
"""
import logging
import os
from datetime import datetime
from typing import Callable
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import ValidationError
from strands import tool, ToolContext

from handlers.error_handler import ErrorHandler
from models.data_models import TransportationExpenseFormInput, GeneralExpenseFormInput

_logger = logging.getLogger(__name__)

# テンプレートパス定数
_TRANSPORTATION_EXPENSE_TEMPLATE = "template/交通費精算申請書テンプレート.xlsx"
_GENERAL_EXPENSE_TEMPLATE = "template/経費精算申請書テンプレート.xlsx"

# Excel 数式インジェクション対策: =, +, -, @ で始まる文字列はセル数式として解釈される
_FORMULA_CHARS = frozenset(("=", "+", "-", "@"))


def _sanitize_cell(value: object) -> object:
    """Excel 数式インジェクションを防ぐため、数式トリガー文字で始まる文字列に
    アポストロフィを付加する。数値・None はそのまま返す。
    """
    if isinstance(value, str) and value and value[0] in _FORMULA_CHARS:
        return "'" + value
    return value


def _generate_form(
    template_path: str,
    applicant_name: str,
    application_date: str,
    validated,
    write_detail_rows: Callable,
    output_filename_prefix: str,
) -> dict:
    """共通フォーム生成処理（テンプレート読み込み → 書き込み → 保存）。

    Args:
        template_path: テンプレートファイルのパス
        applicant_name: 申請者名
        application_date: 申請日
        validated: バリデーション済みの入力データ
        write_detail_rows: 明細行書き込み関数
        output_filename_prefix: 出力ファイル名のプレフィックス

    Returns:
        dict: {"success": bool, "file_path": Optional[str], "error_message": Optional[str]}
    """
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    unique_suffix = uuid4().hex[:8]
    file_path = f"output/{output_filename_prefix}_{timestamp}_{unique_suffix}.xlsx"

    if not os.path.exists(template_path):
        _logger.error(
            "テンプレートファイルが見つかりません: %s, file_path=%s",
            template_path, file_path,
        )
        return {
            "success": False,
            "file_path": None,
            "error_message": "申請書テンプレートが見つかりません。システム管理者にお問い合わせください。",
        }

    try:
        wb = load_workbook(template_path)
        ws = wb.active

        # ヘッダー書き込み
        ws["B3"] = _sanitize_cell(applicant_name)
        ws["B4"] = _sanitize_cell(application_date)

        # 明細行書き込み
        write_detail_rows(ws, validated.items)

        # output/ディレクトリ自動作成
        os.makedirs("output", exist_ok=True)

        wb.save(file_path)
        return {"success": True, "file_path": file_path, "error_message": None}

    except Exception as e:
        _logger.error("申請書保存エラー: file_path=%s", file_path, exc_info=True)
        return {
            "success": False,
            "file_path": None,
            "error_message": ErrorHandler.handle_file_save_error(e),
        }


def _write_transportation_expense_rows(ws, items: list) -> None:
    """交通費精算申請書テンプレートの明細行に区間情報を書き込む。

    Args:
        ws: openpyxlのワークシートオブジェクト
        items: バリデーション済み区間情報リスト
    """
    total_amount = 0
    for i, item in enumerate(items, start=1):
        row = 6 + i  # 行7から開始
        amount = item.get("amount", 0)
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=_sanitize_cell(item.get("travel_date", "")))
        ws.cell(row=row, column=3, value=_sanitize_cell(item.get("departure", "")))
        ws.cell(row=row, column=4, value=_sanitize_cell(item.get("destination", "")))
        ws.cell(row=row, column=5, value=_sanitize_cell(item.get("transport_type", "")))
        ws.cell(row=row, column=6, value=amount)
        ws.cell(row=row, column=7, value=_sanitize_cell(item.get("purpose", "")))
        total_amount += amount

    # 合計金額をH9に書き込み
    ws.cell(row=9, column=8, value=total_amount)


def _write_general_expense_rows(ws, items: list) -> None:
    """経費精算申請書テンプレートの明細行に経費明細を書き込む。

    Args:
        ws: openpyxlのワークシートオブジェクト
        items: バリデーション済み経費明細リスト
    """
    for i, item in enumerate(items, start=1):
        row = 6 + i  # 行7から開始
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=_sanitize_cell(item.get("purchase_date", "")))
        ws.cell(row=row, column=3, value=_sanitize_cell(item.get("store_name", "")))
        ws.cell(row=row, column=4, value=_sanitize_cell(item.get("item_name", "")))
        ws.cell(row=row, column=5, value=_sanitize_cell(item.get("expense_category", "")))
        ws.cell(row=row, column=6, value=item.get("amount", 0))
        ws.cell(row=row, column=7, value=_sanitize_cell(item.get("purpose", "")))


@tool(context=True)
def generate_transportation_expense_form(items: list, tool_context: ToolContext) -> dict:
    """交通費精算申請書を生成する。

    区間情報リストをExcelテンプレートに埋め込み、申請書ファイルを出力する。

    Args:
        items: 区間情報リスト。各要素は以下のフィールドを持つ辞書:
            - travel_date (str): 移動日（YYYY-MM-DD）【必須】
            - departure (str): 出発地【必須】
            - destination (str): 目的地【必須】
            - transport_type (str): 交通手段【必須】
            - amount (int): 金額（円）【必須】
            - purpose (str): 業務目的【必須】

    Returns:
        dict: {"success": bool, "file_path": Optional[str], "error_message": Optional[str]}
    """
    applicant_name = tool_context.invocation_state.get("applicant_name")
    application_date = tool_context.invocation_state.get("application_date")

    if not applicant_name or not application_date:
        return {
            "success": False,
            "file_path": None,
            "error_message": "申請者情報の取得に失敗しました。再度お試しください。",
        }

    _logger.info("交通費精算申請書生成開始: 申請者=%s, 明細件数=%d", applicant_name, len(items))

    try:
        validated = TransportationExpenseFormInput(items=items)
    except ValidationError as e:
        _logger.error("入力バリデーションエラー", exc_info=True)
        return {
            "success": False,
            "file_path": None,
            "error_message": ErrorHandler.handle_validation_error(e),
        }

    result = _generate_form(
        template_path=_TRANSPORTATION_EXPENSE_TEMPLATE,
        applicant_name=applicant_name,
        application_date=application_date,
        validated=validated,
        write_detail_rows=_write_transportation_expense_rows,
        output_filename_prefix="交通費精算申請書",
    )

    if result["success"]:
        _logger.info("交通費精算申請書生成完了: %s", result["file_path"])

    return result


@tool(context=True)
def generate_general_expense_form(items: list, tool_context: ToolContext) -> dict:
    """経費精算申請書を生成する。

    経費明細リストをExcelテンプレートに埋め込み、申請書ファイルを出力する。

    Args:
        items: 経費明細リスト。各要素は以下のフィールドを持つ辞書:
            - purchase_date (str): 購入日（YYYY-MM-DD）【必須】
            - store_name (str): 店舗名【必須】
            - item_name (str): 品目【必須】
            - expense_category (str): 経費区分【必須】
            - amount (int): 金額（円）【必須】
            - purpose (str): 業務目的【必須】

    Returns:
        dict: {"success": bool, "file_path": Optional[str], "error_message": Optional[str]}
    """
    applicant_name = tool_context.invocation_state.get("applicant_name")
    application_date = tool_context.invocation_state.get("application_date")

    if not applicant_name or not application_date:
        return {
            "success": False,
            "file_path": None,
            "error_message": "申請者情報の取得に失敗しました。再度お試しください。",
        }

    _logger.info("経費精算申請書生成開始: 申請者=%s, 明細件数=%d", applicant_name, len(items))

    try:
        validated = GeneralExpenseFormInput(items=items)
    except ValidationError as e:
        _logger.error("入力バリデーションエラー", exc_info=True)
        return {
            "success": False,
            "file_path": None,
            "error_message": ErrorHandler.handle_validation_error(e),
        }

    result = _generate_form(
        template_path=_GENERAL_EXPENSE_TEMPLATE,
        applicant_name=applicant_name,
        application_date=application_date,
        validated=validated,
        write_detail_rows=_write_general_expense_rows,
        output_filename_prefix="経費精算申請書",
    )

    if result["success"]:
        _logger.info("経費精算申請書生成完了: %s", result["file_path"])

    return result
