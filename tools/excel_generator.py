"""
Excel申請書生成ツール

Excel形式の経費精算申請書を生成する。
- 「経費精算申請エージェント(receipt_expense_agent)」：receipt_excel_generatorツールを利用
- 「交通費精算申請エージェント(transportation_expense_agent)」：transportation_excel_generatorツールを利用

"""
import logging
from typing import List, Tuple, Dict
from datetime import datetime
from pathlib import Path
from abc import ABC, abstractmethod
from strands import tool, ToolContext
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill
from pydantic import ValidationError
from handlers.error_handler import ErrorHandler
from models.data_models import RouteInput, InvocationState, ReceiptExpenseInput

logger = logging.getLogger(__name__)


# ==================== クラスの定義 ====================
class ExcelStyleManager:
    """Excelスタイルの定義と適用を管理するクラス。"""
    
    def __init__(self):
        self.styles = self._create_style_definitions()
    
    def _create_style_definitions(self) -> dict:
        """Excel用の共通スタイル定義を作成する。"""
        return {
            "header_font": Font(bold=True, size=12),
            "header_fill": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            "header_alignment": Alignment(horizontal="center", vertical="center"),
            "title_font": Font(bold=True, size=14),
            "title_alignment": Alignment(horizontal="center", vertical="center"),
            "label_font": Font(bold=True, size=12),
            "label_fill": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
            "data_alignment_center": Alignment(horizontal="center", vertical="center"),
            "data_alignment_right": Alignment(horizontal="right", vertical="center"),
            "total_font": Font(bold=True, size=12),
            "total_alignment": Alignment(horizontal="right", vertical="center"),
            "approval_pending_font": Font(color="008000", bold=True),
        }
    
    def apply_header_style(self, cell) -> None:
        """セルにヘッダースタイルを適用する。"""
        cell.font = self.styles["header_font"]
        cell.fill = self.styles["header_fill"]
        cell.alignment = self.styles["header_alignment"]
    
    def apply_title_style(self, cell) -> None:
        """セルにタイトルスタイルを適用する。"""
        cell.font = self.styles["title_font"]
        cell.alignment = self.styles["title_alignment"]
    
    def apply_label_style(self, cell) -> None:
        """セルにラベルスタイルを適用する。"""
        cell.font = self.styles["label_font"]
        cell.fill = self.styles["label_fill"]
        cell.alignment = self.styles["data_alignment_center"]
    
    def apply_data_style(self, cell, alignment: str = "center") -> None:
        """セルにデータスタイルを適用する。"""
        if alignment == "right":
            cell.alignment = self.styles["data_alignment_right"]
        else:
            cell.alignment = self.styles["data_alignment_center"]
    
    def apply_approval_pending_style(self, cell) -> None:
        """セルに承認待ちスタイルを適用する。"""
        cell.font = self.styles["approval_pending_font"]


class ExcelGeneratorBase(ABC):
    """Excel申請書生成の基底クラス"""
    
    COLUMN_WIDTHS = {"A": 12, "B": 15, "C": 15, "D": 15, "E": 12, "F": 15, "G": 12}
    
    def __init__(self, tool_context: ToolContext):
        self.tool_context = tool_context
        self.style_manager = ExcelStyleManager()
        self._error_handler = ErrorHandler()
    
    @abstractmethod
    def generate(self, **kwargs) -> Dict:
        """Excel申請書を生成する抽象メソッド"""
        pass
    
    def _get_applicant_info(self) -> Tuple[str, str]:
        """invocation_stateから申請者名と申請日を取得する。取得できない場合はデフォルト値を返す。"""
        default_name = "未設定"
        default_date = datetime.now().strftime("%Y-%m-%d")
        
        if not self.tool_context or not self.tool_context.invocation_state:
            logger.info("tool_contextまたはinvocation_stateが存在しません。デフォルト値を使用します。")
            return default_name, default_date
        
        try:
            state = InvocationState(**self.tool_context.invocation_state)
            logger.info(f"申請者情報を取得しました: {state.applicant_name}, {state.application_date}")
            return state.applicant_name, state.application_date
        except ValidationError as e:
            logger.error(f"申請者情報の取得に失敗しました: {str(e)}", exc_info=True)
            return default_name, default_date
    
    def _prepare_excel_file(self, prefix: str, title: str) -> Tuple[Workbook, Worksheet, Path]:
        """Excelファイルの準備（ファイル名生成、ディレクトリ確保、ワークブック作成）を行う。"""
        # ファイル名生成
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{prefix}_{timestamp}.xlsx"
        
        # 出力ディレクトリ確保(outputフォルダに出力する)
        try:
            output_path = Path("output")
            output_path.mkdir(parents=True, exist_ok=True)
            logger.info(f"出力ディレクトリを確認しました: {output_path}")
        except OSError as e:
            logger.error(f"出力ディレクトリの作成に失敗しました: {str(e)}", exc_info=True)
            raise
        
        file_path = output_path / filename
        
        # ワークブック作成
        wb = Workbook()
        ws = wb.active
        ws.title = title
        
        return wb, ws, file_path
    
    def _save_workbook(self, wb: Workbook, file_path: Path) -> Tuple[bool, str]:
        """ワークブックをファイルに保存する。"""
        try:
            wb.save(file_path)
            logger.info(f"ファイルを正常に保存しました: {file_path}")
            return True, f"申請書を正常に作成しました: {file_path}"
        except (IOError, PermissionError) as e:
            logger.error(f"ファイルの保存に失敗しました: {str(e)} | file_path: {file_path}", exc_info=True)
            user_message = self._error_handler.handle_file_save_error(e)
            return False, user_message
        except Exception as e:
            logger.error(f"ファイルの保存中に予期しないエラーが発生しました: {str(e)} | file_path: {file_path}", exc_info=True)
            user_message = self._error_handler.handle_file_save_error(e)
            return False, user_message
    
    def _write_title_and_applicant_info(self, ws: Worksheet, title: str, applicant_name: str, application_date: str) -> int:
        """タイトル行と申請者情報を書き込み、次の行番号を返す。"""
        # タイトル行
        ws["A1"] = title
        self.style_manager.apply_title_style(ws["A1"])
        ws.merge_cells("A1:G1")
        
        # 申請者情報
        current_row = 3
        
        # 申請者名
        ws[f"A{current_row}"] = "申請者名"
        self.style_manager.apply_label_style(ws[f"A{current_row}"])
        ws[f"B{current_row}"] = applicant_name
        self.style_manager.apply_data_style(ws[f"B{current_row}"], "center")
        current_row += 1
        
        # 申請日
        ws[f"A{current_row}"] = "申請日"
        self.style_manager.apply_label_style(ws[f"A{current_row}"])
        ws[f"B{current_row}"] = application_date
        self.style_manager.apply_data_style(ws[f"B{current_row}"], "center")
        current_row += 1
        
        # 空行
        current_row += 1
        
        return current_row
    
    def _write_table_header(self, ws: Worksheet, row: int, headers: List[str]) -> int:
        """テーブルヘッダーを書き込み、次の行番号を返す。"""
        columns = ["A", "B", "C", "D", "E", "F", "G"]
        
        for col, header_text in zip(columns, headers):
            cell = ws[f"{col}{row}"]
            cell.value = header_text
            self.style_manager.apply_header_style(cell)
        
        return row + 1
    
    def _write_total_row(self, ws: Worksheet, row: int, label: str, amount: float) -> None:
        """合計行を書き込む。"""
        ws[f"E{row}"] = label
        ws[f"E{row}"].font = self.style_manager.styles["total_font"]
        ws[f"E{row}"].alignment = self.style_manager.styles["total_alignment"]
        
        ws[f"F{row}"] = f"¥{amount:,.0f}"
        ws[f"F{row}"].font = self.style_manager.styles["total_font"]
        ws[f"F{row}"].alignment = self.style_manager.styles["total_alignment"]


class ReceiptExcelGenerator(ExcelGeneratorBase):
    """経費精算申請書を生成するクラス。"""
    
    FILE_PREFIX = "経費精算申請書"
    SHEET_TITLE = "経費精算申請書"
    
    def generate(self, store_name: str, amount: float, date: str, items: List[str], expense_category: str) -> Dict:
        logger.info(
            f"ReceiptExcelGeneratorが呼び出されました: store={store_name}, amount={amount}, date={date}"
        )
        
        try:
            # 申請者情報を取得
            applicant_name, application_date = self._get_applicant_info()
            
            # Pydanticモデルでバリデーション
            try:
                input_data = ReceiptExpenseInput(
                    store_name=store_name,
                    amount=amount,
                    date=date,
                    items=items,
                    expense_category=expense_category
                )
            except ValidationError as e:
                logger.error(f"入力バリデーションエラー（経費精算申請書）: {str(e)}", exc_info=True)
                error_message = self._error_handler.handle_validation_error(e)
                return {
                    "success": False,
                    "file_path": "",
                    "message": error_message
                }
            
            # Excelファイルの準備
            wb, ws, file_path = self._prepare_excel_file(self.FILE_PREFIX, self.SHEET_TITLE)
            
            # 領収書データを書き込み
            self._write_receipt_data(ws, applicant_name, application_date, input_data)
            
            # 列幅の調整
            for col, width in self.COLUMN_WIDTHS.items():
                ws.column_dimensions[col].width = width
            
            # ファイルを保存
            success, message = self._save_workbook(wb, file_path)
            if success:
                return {
                    "success": True,
                    "file_path": str(file_path),
                    "message": message
                }
            else:
                return {
                    "success": False,
                    "file_path": "",
                    "message": message
                }
        
        except Exception as e:
            logger.error(f"経費精算申請書の生成中に予期しないエラーが発生しました: {str(e)}", exc_info=True)
            return {
                "success": False,
                "file_path": "",
                "message": f"エラー: 申請書の生成に失敗しました - {str(e)}"
            }
    
    def _write_receipt_data(self, ws: Worksheet, applicant_name: str, application_date: str, data: ReceiptExpenseInput) -> None:
        """領収書データをワークシートに書き込む。"""
        # タイトル行と申請者情報を書き込み
        current_row = self._write_title_and_applicant_info(ws, self.SHEET_TITLE, applicant_name, application_date)
        
        # テーブルヘッダーを書き込み
        headers = ["No", "購入日", "店舗名", "品目", "経費区分", "金額", "承認状況"]
        current_row = self._write_table_header(ws, current_row, headers)
        
        # データ行（1行のみ）
        ws[f"A{current_row}"] = 1
        self.style_manager.apply_data_style(ws[f"A{current_row}"], "right")
        
        ws[f"B{current_row}"] = data.date
        self.style_manager.apply_data_style(ws[f"B{current_row}"], "center")
        
        ws[f"C{current_row}"] = data.store_name
        self.style_manager.apply_data_style(ws[f"C{current_row}"], "center")
        
        ws[f"D{current_row}"] = ", ".join(data.items)
        self.style_manager.apply_data_style(ws[f"D{current_row}"], "center")
        
        ws[f"E{current_row}"] = data.expense_category
        self.style_manager.apply_data_style(ws[f"E{current_row}"], "center")
        
        ws[f"F{current_row}"] = f"¥{data.amount:,.0f}"
        self.style_manager.apply_data_style(ws[f"F{current_row}"], "right")
        
        ws[f"G{current_row}"] = "承認待ち"
        self.style_manager.apply_approval_pending_style(ws[f"G{current_row}"])
        self.style_manager.apply_data_style(ws[f"G{current_row}"], "center")
        
        current_row += 1
        
        # 合計行
        current_row += 1
        self._write_total_row(ws, current_row, "合計金額", data.amount)


class TransportationExcelGenerator(ExcelGeneratorBase):
    """交通費申請書を生成するクラス。"""
    
    FILE_PREFIX = "交通費申請書"
    SHEET_TITLE = "交通費申請書"
    TRANSPORT_TYPE_MAP = {"train": "電車", "bus": "バス", "taxi": "タクシー", "airplane": "飛行機"}
    
    def generate(self, routes: List[dict]) -> Dict:
        logger.info(
            f"TransportationExcelGeneratorが呼び出されました: routes_count={len(routes) if routes else 0}"
        )
        
        try:
            # 申請者情報を取得
            applicant_name, application_date = self._get_applicant_info()
            
            # データバリデーション
            if not routes:
                logger.error("経路データが空です")
                return {
                    "success": False,
                    "file_path": "",
                    "total_cost": 0,
                    "message": "エラー: 経路データが空です"
                }
            
            # 各経路データをPydanticモデルで検証
            validated_routes = []
            for i, route in enumerate(routes):
                try:
                    validated_route = RouteInput(**route)
                    validated_routes.append(validated_route)
                except ValidationError as e:
                    logger.error(f"経路データのバリデーションエラー（交通費申請書）: route_index={i+1}, {str(e)}", exc_info=True)
                    error_message = self._error_handler.handle_validation_error(e)
                    return {
                        "success": False,
                        "file_path": "",
                        "total_cost": 0,
                        "message": error_message
                    }
            
            # 合計交通費の計算
            total_cost = sum(route.cost for route in validated_routes)
            
            logger.info(
                f"経路データの検証が完了しました。合計交通費: ¥{total_cost:,.0f} ({len(validated_routes)}件)"
            )
            
            # Excelファイルの準備
            wb, ws, file_path = self._prepare_excel_file(self.FILE_PREFIX, self.SHEET_TITLE)
            
            # 経路データを書き込み
            self._write_routes_table(ws, applicant_name, application_date, validated_routes, total_cost)
            
            # 列幅の調整
            for col, width in self.COLUMN_WIDTHS.items():
                ws.column_dimensions[col].width = width
            
            # ファイルを保存
            success, message = self._save_workbook(wb, file_path)
            if success:
                return {
                    "success": True,
                    "file_path": str(file_path),
                    "total_cost": total_cost,
                    "message": message
                }
            else:
                return {
                    "success": False,
                    "file_path": "",
                    "total_cost": 0,
                    "message": message
                }
        
        except Exception as e:
            logger.error(f"交通費申請書の生成中に予期しないエラーが発生しました: {str(e)}", exc_info=True)
            return {
                "success": False,
                "file_path": "",
                "total_cost": 0,
                "message": f"エラー: 申請書の生成に失敗しました - {str(e)}"
            }
    
    def _write_routes_table(self, ws: Worksheet, applicant_name: str, application_date: str, routes: List[RouteInput], total_cost: float) -> None:
        """経路テーブルをワークシートに書き込む。"""
        # タイトル行と申請者情報を書き込み
        current_row = self._write_title_and_applicant_info(ws, self.SHEET_TITLE, applicant_name, application_date)
        
        # テーブルヘッダーを書き込み
        headers = ["No", "日付", "出発地", "目的地", "交通手段", "費用", "承認状況"]
        current_row = self._write_table_header(ws, current_row, headers)
        
        # データ行
        for i, route in enumerate(routes, start=1):
            ws[f"A{current_row}"] = i
            self.style_manager.apply_data_style(ws[f"A{current_row}"], "right")
            
            ws[f"B{current_row}"] = route.date
            self.style_manager.apply_data_style(ws[f"B{current_row}"], "center")
            
            ws[f"C{current_row}"] = route.departure
            self.style_manager.apply_data_style(ws[f"C{current_row}"], "center")
            
            ws[f"D{current_row}"] = route.destination
            self.style_manager.apply_data_style(ws[f"D{current_row}"], "center")
            
            transport_name = self.TRANSPORT_TYPE_MAP.get(route.transport_type, route.transport_type)
            ws[f"E{current_row}"] = transport_name
            self.style_manager.apply_data_style(ws[f"E{current_row}"], "center")
            
            ws[f"F{current_row}"] = f"¥{route.cost:,.0f}"
            self.style_manager.apply_data_style(ws[f"F{current_row}"], "right")
            
            ws[f"G{current_row}"] = "承認待ち"
            self.style_manager.apply_approval_pending_style(ws[f"G{current_row}"])
            self.style_manager.apply_data_style(ws[f"G{current_row}"], "center")
            
            current_row += 1
        
        # 合計行
        current_row += 1
        self._write_total_row(ws, current_row, "合計交通費", total_cost)


#==================== 専門エージェントによるツール関数==================== 
# 経費精算申請エージェント用の申請書作成ツール
@tool(context=True)
def receipt_excel_generator(
    store_name: str,
    amount: float,
    date: str,
    items: List[str],
    expense_category: str,
    tool_context: ToolContext
) -> dict:
    """
    経費精算申請エージェントが利用します。
    経費精算申請書をExcel形式で生成します

    Args:
        store_name: 店舗名
        amount: 金額（円）
        date: 日付（YYYY-MM-DD形式）
        items: 品目のリスト
        expense_category: 経費区分

    Returns:
        dict: {
            "success": bool,         # 成功フラグ
            "file_path": str,        # 保存されたファイルのパス
            "message": str           # 結果メッセージ
        }
    """
    # ReceiptExcelGeneratorクラスを使用して申請書を生成
    generator = ReceiptExcelGenerator(tool_context)
    return generator.generate(
        store_name=store_name,
        amount=amount,
        date=date,
        items=items,
        expense_category=expense_category
    )



# 交通費精算申請エージェント用の申請書作成ツール
@tool(context=True)
def transportation_excel_generator(
    routes: List[dict],
    tool_context: ToolContext
) -> dict:

    """
    交通費精算申請エージェントが利用します
    複数の経路を一覧表形式で表示し、交通費申請書をExcel形式で生成します
    
    Args:
        routes: 経路データのリスト。各要素は以下のキーを持つ辞書:
            - departure (str): 出発地
            - destination (str): 目的地
            - date (str): 日付（YYYY-MM-DD形式）
            - transport_type (str): 交通手段 (train/bus/taxi/airplane)
            - cost (float): 費用
            - notes (str, optional): 備考
    
    Returns:
        dict: {
            "success": bool,         # 成功フラグ
            "file_path": str,        # 保存されたファイルのパス
            "total_cost": float,     # 合計交通費
            "message": str           # 結果メッセージ
        }
    """
    # TransportationExcelGeneratorクラスを使用して申請書を生成
    generator = TransportationExcelGenerator(tool_context)
    return generator.generate(routes=routes)


